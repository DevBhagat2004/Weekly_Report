#!/usr/bin/env python3
"""
Automated Weekly Reports Generator
Processes raw data and generates formatted Excel reports
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime, timedelta
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('weekly_report.log'),
        logging.StreamHandler()
    ]
)

class WeeklyReportGenerator:
    def __init__(self, config_file='config.json'):
        """Initialize the report generator with configuration"""
        self.config = self.load_config(config_file)
        self.data = None
        self.cleaned_data = None
        
    def load_config(self, config_file):
        """Load configuration settings"""
        # Default configuration
        default_config = {
            'input_file': 'raw_data.csv',
            'output_file': f'weekly_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
            'template_file': 'report_template.xlsx',
            'data_columns': ['Date', 'Product', 'Sales', 'Units', 'Revenue', 'Region'],
            'date_column': 'Date',
            'numeric_columns': ['Sales', 'Units', 'Revenue']
        }
        
        # Try to load from file if it exists
        if os.path.exists(config_file):
            import json
            try:
                with open(config_file, 'r') as f:
                    loaded_config = json.load(f)
                default_config.update(loaded_config)
            except Exception as e:
                logging.warning(f"Could not load config file: {e}. Using defaults.")
        
        return default_config
    
    def load_raw_data(self, file_path=None):
        """Load raw data from CSV file"""
        if file_path is None:
            file_path = self.config['input_file']
        
        try:
            logging.info(f"Loading raw data from {file_path}")
            
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    self.data = pd.read_csv(file_path, encoding=encoding)
                    logging.info(f"Successfully loaded data with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if self.data is None:
                raise ValueError("Could not read file with any encoding")
            
            logging.info(f"Loaded {len(self.data)} rows of data")
            return True
            
        except Exception as e:
            logging.error(f"Error loading raw data: {e}")
            return False
    
    def clean_data(self):
        """Clean and process the raw data"""
        try:
            logging.info("Starting data cleaning process")
            self.cleaned_data = self.data.copy()
            
            # Convert date column to datetime
            if self.config['date_column'] in self.cleaned_data.columns:
                self.cleaned_data[self.config['date_column']] = pd.to_datetime(
                    self.cleaned_data[self.config['date_column']], 
                    errors='coerce'
                )
            
            # Clean numeric columns
            for col in self.config['numeric_columns']:
                if col in self.cleaned_data.columns:
                    # Remove currency symbols and commas
                    self.cleaned_data[col] = self.cleaned_data[col].astype(str).str.replace(r'[$,]', '', regex=True)
                    # Convert to numeric
                    self.cleaned_data[col] = pd.to_numeric(self.cleaned_data[col], errors='coerce')
            
            # Remove rows with missing critical data
            critical_columns = [self.config['date_column']] + self.config['numeric_columns']
            initial_rows = len(self.cleaned_data)
            self.cleaned_data = self.cleaned_data.dropna(subset=critical_columns)
            final_rows = len(self.cleaned_data)
            
            logging.info(f"Removed {initial_rows - final_rows} rows with missing data")
            
            # Filter for current week's data
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            if self.config['date_column'] in self.cleaned_data.columns:
                self.cleaned_data = self.cleaned_data[
                    (self.cleaned_data[self.config['date_column']] >= start_date) &
                    (self.cleaned_data[self.config['date_column']] <= end_date)
                ]
            
            logging.info(f"Data cleaning completed. Final dataset: {len(self.cleaned_data)} rows")
            return True
            
        except Exception as e:
            logging.error(f"Error during data cleaning: {e}")
            return False
    
    def generate_summary_stats(self):
        """Generate summary statistics"""
        try:
            summary = {}
            
            for col in self.config['numeric_columns']:
                if col in self.cleaned_data.columns:
                    summary[f'Total_{col}'] = self.cleaned_data[col].sum()
                    summary[f'Avg_{col}'] = self.cleaned_data[col].mean()
                    summary[f'Max_{col}'] = self.cleaned_data[col].max()
                    summary[f'Min_{col}'] = self.cleaned_data[col].min()
            
            # Additional metrics
            summary['Total_Records'] = len(self.cleaned_data)
            summary['Report_Date'] = datetime.now().strftime("%Y-%m-%d")
            
            # Group by categorical columns for insights
            if 'Product' in self.cleaned_data.columns and 'Revenue' in self.cleaned_data.columns:
                top_products = self.cleaned_data.groupby('Product')['Revenue'].sum().sort_values(ascending=False)
                summary['Top_Product'] = top_products.index[0] if len(top_products) > 0 else 'N/A'
                summary['Top_Product_Revenue'] = top_products.iloc[0] if len(top_products) > 0 else 0
            
            if 'Region' in self.cleaned_data.columns and 'Revenue' in self.cleaned_data.columns:
                top_regions = self.cleaned_data.groupby('Region')['Revenue'].sum().sort_values(ascending=False)
                summary['Top_Region'] = top_regions.index[0] if len(top_regions) > 0 else 'N/A'
                summary['Top_Region_Revenue'] = top_regions.iloc[0] if len(top_regions) > 0 else 0
            
            return summary
            
        except Exception as e:
            logging.error(f"Error generating summary stats: {e}")
            return {}
    
    def create_excel_report(self, output_file=None):
        """Create formatted Excel report"""
        if output_file is None:
            output_file = self.config['output_file']
        
        try:
            logging.info(f"Creating Excel report: {output_file}")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create Summary sheet
            summary_sheet = wb.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet)
            
            # Create Data sheet
            data_sheet = wb.create_sheet("Raw Data")
            self._create_data_sheet(data_sheet)
            
            # Create Charts sheet
            charts_sheet = wb.create_sheet("Charts")
            self._create_charts_sheet(charts_sheet)
            
            # Set Summary as active sheet
            wb.active = summary_sheet
            
            # Save workbook
            wb.save(output_file)
            logging.info(f"Excel report saved successfully: {output_file}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating Excel report: {e}")
            return False
    
    def _create_summary_sheet(self, sheet):
        """Create the summary sheet with key metrics"""
        # Title
        sheet['A1'] = "Weekly Report Summary"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        # Date
        sheet['A3'] = "Report Generated:"
        sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Summary statistics
        summary = self.generate_summary_stats()
        row = 5
        
        for key, value in summary.items():
            sheet[f'A{row}'] = key.replace('_', ' ').title()
            sheet[f'B{row}'] = value
            if isinstance(value, (int, float)) and 'Revenue' in key:
                sheet[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Format headers
        for row in range(5, row):
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_data_sheet(self, sheet):
        """Create the raw data sheet"""
        # Add headers
        headers = list(self.cleaned_data.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add data
        for row_idx, row_data in enumerate(dataframe_to_rows(self.cleaned_data, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numbers
                if col_idx in [i+1 for i, col in enumerate(headers) if col in self.config['numeric_columns']]:
                    if 'Revenue' in headers[col_idx-1]:
                        cell.number_format = '$#,##0.00'
                    else:
                        cell.number_format = '#,##0'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_charts_sheet(self, sheet):
        """Create charts and visualizations"""
        sheet['A1'] = "Data Visualizations"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Create product revenue chart if applicable
        if 'Product' in self.cleaned_data.columns and 'Revenue' in self.cleaned_data.columns:
            product_data = self.cleaned_data.groupby('Product')['Revenue'].sum().sort_values(ascending=False).head(10)
            
            # Add chart data
            row = 3
            sheet[f'A{row}'] = "Product"
            sheet[f'B{row}'] = "Revenue"
            
            for idx, (product, revenue) in enumerate(product_data.items(), row + 1):
                sheet[f'A{idx}'] = product
                sheet[f'B{idx}'] = revenue
            
            # Create chart
            chart = BarChart()
            chart.title = "Top Products by Revenue"
            chart.x_axis.title = "Products"
            chart.y_axis.title = "Revenue"
            
            data_ref = Reference(sheet, min_col=2, min_row=row, max_col=2, max_row=row + len(product_data))
            cats_ref = Reference(sheet, min_col=1, min_row=row + 1, max_row=row + len(product_data))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            sheet.add_chart(chart, "D3")
    
    def run_full_process(self):
        """Run the complete report generation process"""
        logging.info("Starting automated weekly report generation")
        
        # Load data
        if not self.load_raw_data():
            return False
        
        # Clean data
        if not self.clean_data():
            return False
        
        # Generate report
        if not self.create_excel_report():
            return False
        
        logging.info("Weekly report generation completed successfully")
        return True

def main():
    """Main function to run the report generator"""
    try:
        # Create sample data if raw data doesn't exist
        if not os.path.exists('raw_data.csv'):
            create_sample_data()
        
        # Initialize and run report generator
        generator = WeeklyReportGenerator()
        success = generator.run_full_process()
        
        if success:
            print("✅ Weekly report generated successfully!")
            print(f"📊 Report saved as: {generator.config['output_file']}")
        else:
            print("❌ Report generation failed. Check logs for details.")
            
    except Exception as e:
        logging.error(f"Unexpected error in main: {e}")
        print(f"❌ Unexpected error: {e}")

def create_sample_data():
    """Create sample data for demonstration"""
    import random
    from datetime import datetime, timedelta
    
    # Generate sample data
    products = ['Product A', 'Product B', 'Product C', 'Product D', 'Product E']
    regions = ['North', 'South', 'East', 'West', 'Central']
    
    data = []
    start_date = datetime.now() - timedelta(days=30)
    
    for i in range(200):
        date = start_date + timedelta(days=random.randint(0, 30))
        product = random.choice(products)
        region = random.choice(regions)
        units = random.randint(1, 100)
        price = random.uniform(10, 500)
        revenue = units * price
        sales = random.randint(1, 20)
        
        data.append({
            'Date': date.strftime('%Y-%m-%d'),
            'Product': product,
            'Sales': sales,
            'Units': units,
            'Revenue': round(revenue, 2),
            'Region': region
        })
    
    # Save sample data
    df = pd.DataFrame(data)
    df.to_csv('raw_data.csv', index=False)
    logging.info("Sample data created: raw_data.csv")

if __name__ == "__main__":
    main()
